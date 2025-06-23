# defender_report/reporting.py
import datetime
import logging
import os
from typing import Dict, List

import pandas as pd
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from defender_report.utils import Spinner

logger = logging.getLogger(__name__)


def _nested_report_folder(
    root_directory: str, department: str, report_date: datetime.date
) -> str:
    """
    Build and return a nested folder path:
      root_directory/
        department/
          {financial_year}/        # e.g. "2024-2025"
            Q{1–4}/                 # fiscal quarter (Apr–Jun=Q1 … Jan–Mar=Q4)
              {MonthName}/          # e.g. "June"
                {YYYY-MM-DD}/       # e.g. "2024-06-15"
    Creates folders if they do not already exist.
    """
    month = report_date.month
    year = report_date.year

    # Calculate financial year
    if month >= 4:
        fy_start = year
    else:
        fy_start = year - 1
    fy_end = fy_start + 1
    financial_year = f"{fy_start}-{fy_end}"

    # Calculate fiscal quarter
    if 4 <= month <= 6:
        quarter = "Q1"
    elif 7 <= month <= 9:
        quarter = "Q2"
    elif 10 <= month <= 12:
        quarter = "Q3"
    else:
        quarter = "Q4"

    month_name = report_date.strftime("%B")
    date_folder = report_date.isoformat()

    path_parts = [
        root_directory,
        department,
        financial_year,
        quarter,
        month_name,
        date_folder,
    ]
    full_path = os.path.join(*path_parts)
    os.makedirs(full_path, exist_ok=True)
    return full_path


def _create_compliance_formats(workbook):
    """
    Return three XlsxWriter formats for the 80% baseline:
      green  = '#00b050', yellow = '#ffff00', red = '#ff0000'
    """
    green = workbook.add_format({"bg_color": "#00b050", "num_format": "0.0%"})
    yellow = workbook.add_format({"bg_color": "#ffff00", "num_format": "0.0%"})
    red = workbook.add_format({"bg_color": "#ff0000", "num_format": "0.0%"})
    return green, yellow, red


def _write_table(
    writer: pd.ExcelWriter,
    dataframe: pd.DataFrame,
    sheet_name: str,
    style_name: str = "Table Style Medium 16",
) -> None:
    """
    Write `dataframe` to `sheet_name` as a simple Excel table without totals.
    """
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    rows, cols = dataframe.shape
    worksheet.add_table(
        0,
        0,
        rows,
        cols - 1,
        {
            "style": style_name,
            "columns": [{"header": c} for c in dataframe.columns],
        },
    )


def _write_summary_table(
    writer: pd.ExcelWriter,
    summary_dataframe: pd.DataFrame,
    sheet_name: str,
    report_date: datetime.date,
) -> None:
    """
    Create a Summary sheet:
      1) Put report_date (e.g. '5-Jun') in A1
      2) Write `summary_dataframe` as 'Table Style Medium 16' at row 2
         with a total row (sum for numeric, average for Compliance)
      3) Apply 80% baseline conditional formatting on Compliance
      4) Append a Baseline legend beneath the table
    """
    workbook = writer.book
    green, yellow, red = _create_compliance_formats(workbook)

    # 1) Create or overwrite the sheet
    worksheet = workbook.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = worksheet
    last_column_index = len(summary_dataframe.columns) - 1

    # 1a) Write the date header in cell A1
    date_format = workbook.add_format({"bold": True, "align": "left", "font_size": 12})
    raw_label  = report_date.strftime("%d-%b")   # e.g. "05-Jun"
    date_label = raw_label.lstrip("0")           # -> "5-Jun"
    worksheet.write(0, 0, date_label, date_format)

    # 2) Write the DataFrame starting at row 2 (0-indexed row 1)
    start_row = 1
    summary_dataframe.to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_row
    )

    # 2a) Convert that range into a Table with a total row
    num_rows, num_cols = summary_dataframe.shape
    table_columns = []
    for column in summary_dataframe.columns:
        if column == "Department":
            table_columns.append({"header": column})
        elif column == "Compliance":
            table_columns.append({"header": column, "total_function": "average"})
        else:
            table_columns.append({"header": column, "total_function": "sum"})

    worksheet.add_table(
        start_row,
        0,
        start_row + num_rows,
        num_cols - 1,
        {
            "style": "Table Style Medium 16",
            "total_row": True,
            "columns": table_columns,
        },
    )

    # 3) Apply conditional formatting on the data rows of Compliance
    comp_index = summary_dataframe.columns.get_loc("Compliance") + 1
    column_letter = get_column_letter(comp_index)
    data_row_start = start_row + 1
    data_row_end = start_row + num_rows
    cf_range = f"{column_letter}{data_row_start + 1}:{column_letter}{data_row_end + 1}"

    worksheet.conditional_format(
        cf_range, {"type": "cell", "criteria": ">=", "value": 0.8, "format": green}
    )
    worksheet.conditional_format(
        cf_range,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.7000001,
            "maximum": 0.8,
            "format": yellow,
        },
    )
    worksheet.conditional_format(
        cf_range, {"type": "cell", "criteria": "<", "value": 0.7, "format": red}
    )

    # 4) Add Baseline legend beneath the table
    legend_items = [
        ("Baseline 80%", None),
        ("> 80% (green)", "#00b050"),
        ("< 80% (yellow)", "#ffff00"),
        ("< 70% (red)", "#ff0000"),
    ]
    legend_start_row = data_row_end + 2
    for offset, (text, color) in enumerate(legend_items):
        fmt_props = {"bold": True, "align": "left"}
        if color:
            fmt_props["bg_color"] = color
        fmt = workbook.add_format(fmt_props)
        worksheet.merge_range(
            legend_start_row + offset,
            0,
            legend_start_row + offset,
            last_column_index,
            text,
            fmt,
        )


def write_full_report(
    all_sheets: Dict[str, pd.DataFrame],
    summary_df: pd.DataFrame,
    sheet_order: List[str],
    output_path: str,
) -> None:
    import datetime

    if "ungrouped" not in sheet_order:
        sheet_order = sheet_order + ["ungrouped"]

    with Spinner("Writing master report"):
        with pd.ExcelWriter(
            output_path, engine="xlsxwriter", datetime_format="yyyy-mm-dd"
        ) as writer:
            workbook = writer.book

            # your brand-new colors
            fmt_g = workbook.add_format({"bg_color": "#00b050", "num_format": "0.0%"})
            fmt_y = workbook.add_format({"bg_color": "#ffff00", "num_format": "0.0%"})
            fmt_r = workbook.add_format({"bg_color": "#ff0000", "num_format": "0.0%"})
            style = {"style": "Table Style Medium 16"}

            # ── 1) write each department sheet exactly as before ───────
            for dept in tqdm(sheet_order, desc="Sheets → master", unit="sheet"):
                df_dept = all_sheets.get(dept, pd.DataFrame())
                df_dept.to_excel(writer, sheet_name=dept, index=False)
                ws = writer.sheets[dept]
                rows, cols = df_dept.shape
                ws.add_table(
                    0,
                    0,
                    rows,
                    cols - 1,
                    {**style, "columns": [{"header": c} for c in df_dept.columns]},
                )

            # ── 2) Summary sheet ───────────────────────────────────────
            ws_sum = workbook.add_worksheet("Summary")
            writer.sheets["Summary"] = ws_sum

            # 2a) date header in row 1, merged A→last column
            report_date = datetime.date.today()
            raw = report_date.strftime("%d-%b")  # “05-Jun” on any OS
            date_str    = raw.lstrip("0")               # “5-Jun”
            last_col = len(summary_df.columns) - 1
            hdr_fmt = workbook.add_format(
                {
                    "bold": True,
                    "align": "center",
                    "font_size": 12,
                    "bg_color": "#00b0f0"
                }
            )
            ws_sum.merge_range(0, 0, 0, last_col, date_str, hdr_fmt)

            # 2b) dump DataFrame starting at row 2 (Excel row index 1)
            summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
            sr, sc = summary_df.shape

            # 2c) add table exactly as old code, but shifted down one row
            ws_sum.add_table(
                1,  # header row at row=1
                0,  # first column A
                1
                + sr,  # include one extra row (total row placeholder, though you don't show totals)
                sc - 1,  # last column index
                {
                    **style,
                    "columns": [{"header": c} for c in summary_df.columns],
                },
            )

            # 2d) conditional formatting on the **data rows** of Compliance
            comp_idx = summary_df.columns.get_loc("Compliance") + 1
            col_letter = get_column_letter(comp_idx)
            # data rows in Excel run from row 2 (first data) to row 1+sr
            rng = f"{col_letter}2:{col_letter}{1 + sr}"
            ws_sum.conditional_format(
                rng, {"type": "cell", "criteria": ">=", "value": 0.8, "format": fmt_g}
            )
            ws_sum.conditional_format(
                rng,
                {
                    "type": "cell",
                    "criteria": "between",
                    "minimum": 0.7000001,
                    "maximum": 0.8,
                    "format": fmt_y,
                },
            )
            ws_sum.conditional_format(
                rng, {"type": "cell", "criteria": "<=", "value": 0.7, "format": fmt_r}
            )

            # 2e) legend under the table (two rows below last data row)
            legend = [
                ("Baseline 80%", None),
                ("> 80% (green)", "#00b050"),
                ("< 80% (yellow)", "#ffff00"),
                ("< 70% (red)", "#ff0000"),
            ]
            legend_start = 1 + sr + 2
            for i, (text, color) in enumerate(legend):
                fmt_cfg = {"bold": True, "align": "left"}
                if color:
                    fmt_cfg["bg_color"] = color
                fmt = workbook.add_format(fmt_cfg)
                ws_sum.write(legend_start + i, 0, text, fmt)

        logger.info("Master report written to %s", output_path)


def write_department_reports(
    all_sheets: Dict[str, pd.DataFrame],
    summary_dataframe: pd.DataFrame,
    sheet_order: List[str],
    output_root: str,
) -> None:
    """
    Generate one report per department in a nested folder tree,
    using the same summary‐sheet helper for identical styling.
    """
    report_date = datetime.date.today()
    departments = list(sheet_order)
    if "ungrouped" not in departments:
        departments.append("ungrouped")

    for department_code in tqdm(departments, desc="Writing per-dept", unit="dept"):
        target_directory = _nested_report_folder(
            output_root, department_code, report_date
        )
        file_name = f"{department_code}_Report_{report_date.isoformat()}.xlsx"
        full_path = os.path.join(target_directory, file_name)

        with Spinner(f"Writing {department_code} report"):
            with pd.ExcelWriter(
                full_path, engine="xlsxwriter", datetime_format="yyyy-mm-dd"
            ) as writer:
                # detail sheet
                _write_table(
                    writer,
                    all_sheets.get(department_code, pd.DataFrame()),
                    department_code,
                )
                # identical summary sheet
                _write_summary_table(writer, summary_dataframe, "Summary", report_date)

        print(f"→ Wrote {department_code} report to {full_path}")
